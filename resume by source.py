#resume by source
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import xlsxwriter
import json
from operator import index
from tabnanny import filename_only
from typing import final
import os
import warnings

#import excel
infile = r"C:\Users\rachellee\Desktop\Resumes\Analysis\spanned_annotations.xlsx"
df = pd.read_excel(infile)
#print(df)
source=df['Source'].unique()
#print(source)
src=df['Source']
#print(src)


s1=df[src=='Kaggle (1)']
s2=df[src=='Kaggle (2)']
s3=df[src=='Kaggle (3)']
s4=df[src=='MITRE']
'''
print('Kaggle (1)')
print(s1)
print('Kaggle (2)')
print(s2)
print('Kaggle (3)')
print(s3)
print('MITRE')
print(s4)
'''


outfile=r"C:\Users\rachellee\Desktop\Resumes\Analysis\spannedbysource.xlsx"

'''
with pd.ExcelWriter(outfile, engine='openpyxl', mode='a') as writer:
    s1.to_excel(writer, sheet_name='Kaggle (1)')
    s2.to_excel(writer, sheet_name='Kaggle (2)')
    s3.to_excel(writer, sheet_name='Kaggle (3)')
    s4.to_excel(writer, sheet_name='MITRE')
'''

wb=openpyxl.load_workbook(outfile)
sheet=wb.sheetnames
    

workBook=load_workbook(outfile)
sheets=workBook.sheetnames
i=1
for s_name in sheets:
    print(s_name)
    df2=pd.read_excel(outfile, sheet_name=s_name)

    print(df2)

    #get total counts for each label for each source
    
    
    count2=df2['Label'].value_counts()
    #print(count2)
    outdata=r"C:\Users\rachellee\Desktop\Resumes\Analysis\spanneddatabysource.xlsx"

    with pd.ExcelWriter(outdata, engine='openpyxl', mode='a') as writer:
        count2.to_excel(writer, sheet_name=s_name)


#count=df2['Filename'].value_counts()
#print(count)