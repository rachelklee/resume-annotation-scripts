from fnmatch import fnmatch
import sndhdr
import pandas as pd
import numpy as np
import openpyxl
import xlsxwriter
import xlrd 

fn=r"C:\Users\rachellee\Desktop\Resumes\Analysis\stats.xlsx"
wb=openpyxl.load_workbook(fn)

#process excel file
sheet='Annotations per Resume'
wb.active=wb[sheet]
sh=wb.active
df=pd.read_excel(fn)


 #calculate average
col=sh.cell(row=1,column=2)
coltitle=col.value

avgval=df[coltitle].mean()
print(avgval)

#print average to excel
count=0
for row in sh:
    if not all([cell.value is None for cell in row]):
        count+=1
newrow=count+1

cell="A"+str(newrow)
print(cell)

sh[cell]='Mean'
sh.cell(row=newrow,column=2).value=avgval
wb.save(fn)

    