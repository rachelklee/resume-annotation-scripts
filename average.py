from fnmatch import fnmatch
import sndhdr
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import xlrd 
import xlwt
from openpyxl import Workbook

#fn=r"C:\Users\rachellee\Desktop\Resumes\Analysis\testing.xlsx"
fn=r"C:\Users\rachellee\Desktop\Resumes\Analysis\stats.xlsx"
#wb=openpyxl.load_workbook(fn)

workBook=load_workbook(fn)
sheets=workBook.sheetnames
i=1
for s_name in sheets:
    print(s_name)
    sheet=workBook[s_name]
    #wb.active=wb[sheet]
    #sh=wb.active
    xls=pd.ExcelFile(fn)
    df=pd.read_excel(xls, s_name)
    
    colname=df.columns[1]

    avg=df[colname].mean()
    print(avg)

    count=0
    for row in sheet:
        if not all([cell.value is None for cell in row]):
            count+=1
    newrow=count+1

    cell="A"+str(newrow)
    print(cell)

    sheet[cell]='Mean'
    sheet.cell(row=newrow,column=2).value=avg
    workBook.save(fn)

    #sheet[cell]='Mean'
    #sheet.cell(row=newrow,column=2).value=avg
    #wb.save(fn)
